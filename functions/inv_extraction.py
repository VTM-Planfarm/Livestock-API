import os, openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from functions.vars import annual_stock_class_data
from copy import deepcopy

def extract_inventories_from_excel(inventory_sheet: openpyxl.Workbook, livestock: str) -> list:
    seasonal_data = []

    seasonal_data.append(extract_seasonal_data(inventory_sheet, livestock))

    return seasonal_data


def extract_seasonal_data(
        inventory_sheet: openpyxl.Workbook,
        livestock: str
) -> dict:
    seasonal_sheet = inventory_sheet[f"{livestock}SeasonalData"]
    stock_data = {}

    for col in range(3, 19):
        stock_class = seasonal_sheet.cell(2, col).value
        stock_data[stock_class] = deepcopy(annual_stock_class_data)

        for row in range(3, 34):
            key = seasonal_sheet.cell(row, 2).value
            season = seasonal_sheet.cell(row, 1).value
            value = seasonal_sheet.cell(row, col).value

            if key in ["crudeProtein", "dryMatterDigestibility", "feedAvailability"]:
                if value == 0:
                    continue

            if season is not None:
                stock_data[stock_class][season.lower()][key] = value
            elif key == "head" or key == "purchaseWeight":
                stock_data[stock_class]["purchases"][0][key] = value
            else:
                stock_data[stock_class][key] = value

    return stock_data


def extract_annual_data(inventory_sheet: openpyxl.Workbook, json_data: dict, livestock: str) -> dict:
    if livestock.lower() == "sheep":
        row = 2
    elif livestock.lower() == "cattle":
        row = 3
    else:
        raise ValueError("Unsupported livestock type")
    
    annual_sheet = inventory_sheet["Annual Data"]

    json_data = extract_lime_data(json_data, annual_sheet, row)
    json_data = extract_fertiliser_data(json_data, annual_sheet, row)
    json_data = extract_fuel_data(json_data, annual_sheet, row)
    json_data = extract_electricity_data(json_data, annual_sheet, inventory_sheet, row)
    json_data = extract_supplementation_data(json_data, annual_sheet, row)
    json_data = extract_feed_data(json_data, annual_sheet, row)
    json_data = extract_chemical_data(json_data, annual_sheet, row)
    json_data = extract_merino_pct(json_data, annual_sheet, row)
    json_data = extract_ewesLambing_rate(json_data, annual_sheet, row)
    json_data = extract_seasonalLambing_rate(json_data, annual_sheet, row)
    json_data = extract_vegetation_data(json_data, inventory_sheet["Vegetation"])

    return json_data


def extract_lime_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["limestone"] = annual_sheet.cell(row, 2).value
    json_data["sheep"][group]["limestoneFraction"] = annual_sheet.cell(row, 3).value

    return json_data


def extract_fertiliser_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["fertiliser"] = {
        "singleSuperphosphate": annual_sheet.cell(row, 4).value,
        "pastureDryland": annual_sheet.cell(row, 5).value, # Urea
        "pastureIrrigated": 0, # Urea
        "cropsDryland": annual_sheet.cell(row, 6).value, # Urea
        "cropsIrrigated": 0, # Urea
        "otherFertilisers": []
    }

    for col in range(7, 21):
        json_data["sheep"][group]["fertiliser"]["otherFertilisers"].append(
            {
                "otherType": annual_sheet.cell(1, col).value,
                "otherDryland": annual_sheet.cell(row, col).value,
                "otherIrrigated": 0  # Assuming no irrigated data for other fertilisers
            }
        )

    return json_data


def extract_fuel_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["diesel"] = annual_sheet.cell(row, 21).value
    
    json_data["sheep"][group]["petrol"] = annual_sheet.cell(row, 22).value
    
    json_data["sheep"][group]["lpg"] = annual_sheet.cell(row, 23).value

    return json_data


def extract_electricity_data(
        json_data: dict, 
        annual_sheet: openpyxl.worksheet.worksheet.Worksheet,
        inventory_sheet: openpyxl.Workbook, 
        row: int,
        group: int = 0
) -> dict:
    json_data["sheep"][group]["electricitySource"] = inventory_sheet["Client detail"].cell(54, 7).value

    if json_data["sheep"][group]["electricitySource"] != "Renewable":
        json_data["sheep"][group]["electricityRenewable"] = annual_sheet.cell(row, 30).value

    json_data["sheep"][group]["electricityUse"] = annual_sheet.cell(row, 31).value

    return json_data


def extract_supplementation_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["mineralSupplementation"] = {
        "mineralBlock": annual_sheet.cell(row, 24).value,
        "mineralBlockUrea": annual_sheet.cell(row, 25).value,
        "weanerBlock": annual_sheet.cell(row, 26).value,
        "weanerBlockUrea": annual_sheet.cell(row, 27).value,
        "dryseasonMix": annual_sheet.cell(row, 28).value,
        "dryseasonMixUrea": annual_sheet.cell(row, 29).value,
    }

    return json_data


def extract_feed_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["grainFeed"] = annual_sheet.cell(row, 32).value

    json_data["sheep"][group]["hayFeed"] = annual_sheet.cell(row, 33).value

    return json_data


def extract_chemical_data(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["herbicide"] = annual_sheet.cell(row, 34).value

    json_data["sheep"][group]["herbicideOther"] = annual_sheet.cell(row, 35).value

    json_data["sheep"][group]["merionoPercent"] = annual_sheet.cell(row, 36).value

    return json_data

def extract_merino_pct(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["merinoPercent"] = annual_sheet.cell(row, 36).value

    return json_data

def extract_ewesLambing_rate(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["ewesLambing"] = {
        "autumn": 0,
        "winter": 0,
        "spring": 0,
        "summer": 0
    }

    season = annual_sheet.cell(row, 37).value
    rate = annual_sheet.cell(row, 38).value

    json_data["sheep"][group]["ewesLambing"][season.lower()] = rate # type: ignore

    return json_data

def extract_seasonalLambing_rate(json_data: dict, annual_sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, group: int = 0) -> dict:
    json_data["sheep"][group]["seasonalLambing"] = {
        "autumn": 0,
        "winter": 0,
        "spring": 0,
        "summer": 0
    }

    season = annual_sheet.cell(row, 39).value
    rate = annual_sheet.cell(row, 40).value

    json_data["sheep"][group]["seasonalLambing"][season.lower()] = rate # type: ignore

    return json_data

def extract_vegetation_data(json_data: dict, vegetation_sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    json_data["vegetation"] = []

    row = 2
    if vegetation_sheet.cell(row, 6).value is None:
        json_data["vegetation"].append(
            {
                "vegetation": {
                    "region": "South West",
                    "treeSpecies": "No tree data available",
                    "soil": "No Soil / Tree data available",
                    "area": 0,
                    "age": 0
                },
                "sheepProportion": [1]
            }
        )
        return json_data
    
    while vegetation_sheet.cell(row, 6).value is not None:
        json_data["vegetation"].append(
            {
                "vegetation": {
                    "region": vegetation_sheet.cell(row, 2).value,
                    "treeSpecies": vegetation_sheet.cell(row, 3).value,
                    "soil": vegetation_sheet.cell(row, 5).value,
                    "area": vegetation_sheet.cell(row, 7).value,
                    "age": vegetation_sheet.cell(row, 8).value
                },
                "sheepProportion": [1]
            }
        )
        row += 1

    return json_data